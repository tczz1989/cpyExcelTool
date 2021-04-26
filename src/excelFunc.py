import xlrd
import xlwt
from openpyxl import load_workbook
from openpyxl import Workbook
# from datetime import date, datetime


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        pass
    return False


def open_excel(filename):
    # 打开文件
    workbook = xlrd.open_workbook(r'../input/'+filename)
    # 获取所有sheet
    # print(workbook.sheet_names())  # [u'sheet1', u'sheet2']
    # sheet2_name = workbook.sheet_names()[1]
    return workbook


def open_sheet(workbook, *sheet_index):
    if sheet_index is None or sheet_index is () or sheet_index is [] or sheet_index is "":
        sheet = workbook.sheet_by_index(0)
    else:
        if is_number(sheet_index):
            # 根据sheet索引或者名称获取sheet内容
            sheet = workbook.sheet_by_index(sheet_index)  # sheet索引从0开始
        else:
            sheet = workbook.sheet_by_name(sheet_index)
    # sheet的名称，行数，列数
    print(sheet.name, sheet.nrows, sheet.ncols)
    return sheet


def read_excel(sheet, rows, cols):
    data = []
    for row in rows:
        for col in cols:
            if len(row) is 1 and len(col) is 1:
                # 参数对应 行, 列, 值
                data.append(sheet.cell_value(row, col))
                # 保存
            elif len(row) is 2 and len(col) is 2:
                for r in range(row[0], row[1]+1):
                    for c in range(col[0], col[1]+1):
                        data.append(sheet.cell_value(r, c))
            elif len(row) is 2 and len(col) is 1:
                for r in range(row[0], row[1]+1):
                    data.append(sheet.cell_value(r, col[0]))
            elif len(row) is 1 and len(col) is 2:
                for c in range(col[0], col[1]+1):
                    data.append(sheet.cell_value(row[0], c))
            else:
                print("row col input error!")
                return
    # # 获取整行和整列的值（数组）
    # rows = sheet.row_values(3)  # 获取第四行内容
    # cols = sheet.col_values(2)  # 获取第三列内容
    # print(rows)
    # print(cols)
    #
    # # 获取单元格内容
    # print(sheet.cell(1, 0).value)
    # print(sheet.cell_value(1, 0))
    # print(sheet.row(1)[0].value)
    #
    # # 获取单元格内容的数据类型
    # print(sheet.cell(1, 0).ctype)
    # print(data)
    return data


def write_excel_xls(output_filename, input_data, output_rows, output_cols, *sheet):
    # # 创建一个workbook 设置编码
    workbook = xlwt.Workbook(encoding='ascii')
    if sheet is None or sheet is () or sheet is [] or sheet is "":
        # 创建一个worksheet
        sheet = workbook.add_sheet('Sheet1')
    # 写入excel
    try:
        index = 0
        for i, row in enumerate(output_rows):
            for col in output_cols:
                if len(row) is 1 and len(col) is 1:
                    # 参数对应 行, 列, 值
                    sheet.write(row, col, label=input_data[index])
                    index += 1
                    # 保存
                elif len(row) is 2 and len(col) is 2:
                    for r in range(row[0], row[1]+1):
                        for c in range(col[0], col[1]+1):
                            sheet.write(r, c, label=input_data[index])
                            index += 1
                elif len(row) is 2 and len(col) is 1:
                    for r in range(row[0], row[1]+1):
                        sheet.write(r, col[0], label=input_data[index])
                        index += 1
                elif len(row) is 1 and len(col) is 2:
                    for c in range(col[0], col[1]+1):
                        sheet.write(row[0], c, label=input_data[index])
                        index += 1
                else:
                    print("row col input error!")
                    return
            workbook.save('../output/'+output_filename)
    # except PermissionError:
    #     print("请关闭写入的目标文件！")
    except Exception as e:
        print(str(e))
    return


def create_excel_xlsx(output_config):
    workbook = Workbook()
    sheet_name = workbook.get_sheet_names()
    workbook.create_sheet(output_config[1][1])
    worksheet = workbook.get_sheet_by_name(sheet_name[0])
    workbook.remove(worksheet)
    workbook.save("../output/" + output_config[0][1])


def write_excel_xlsx(output_filename, input_data, output_rows, output_cols, sheet_name):
    # Create a workbook and add a worksheet.
    workbook = load_workbook("../output/" + output_filename)
    # worksheet = workbook.add_worksheet()
    if sheet_name in workbook.get_sheet_names():
        sheet = workbook.get_sheet_by_name(sheet_name)
    else:
        sheet = workbook.create_sheet(sheet_name)
    # 写入excel
    try:
        index = 0
        for i, row in enumerate(output_rows):
            for col in output_cols:
                if len(row) is 1 and len(col) is 1:
                    # 参数对应 行, 列, 值
                    sheet.cell(row+1, col+1, label=input_data[index])
                    index += 1
                    # 保存
                elif len(row) is 2 and len(col) is 2:
                    for r in range(row[0], row[1]+1):
                        for c in range(col[0], col[1]+1):
                            sheet.cell(r+1, c+1, input_data[index])
                            index += 1
                elif len(row) is 2 and len(col) is 1:
                    for r in range(row[0], row[1]+1):
                        sheet.cell(r+1, col[0]+1, input_data[index])
                        index += 1
                elif len(row) is 1 and len(col) is 2:
                    for c in range(col[0], col[1]+1):
                        sheet.cell(row[0]+1, c+1, input_data[index])
                        index += 1
                else:
                    print("row col input error!")
                    return
            workbook.save('../output/'+output_filename)
    # except PermissionError:
    #     print("请关闭写入的目标文件！")
    except Exception as e:
        print(str(e))
    return


def get_col_name(input_str):
    sections = input_str.split(",")
    col_ranges = []
    for section in sections:
        col_range_str = section.split("-")
        col_range_num = []
        for col in col_range_str:
            col_range_num.append(ord(col.lower().strip())-97)
        col_ranges.append(col_range_num)
    print(col_ranges)
    return col_ranges


def get_row_name(input_str):
    sections = input_str.split(",")
    row_ranges = []
    for section in sections:
        row_range_str = section.split("-")
        row_range_num = []
        for row in row_range_str:
            row_range_num.append(int(row.strip())-1)
        row_ranges.append(row_range_num)
    print(row_ranges)
    return row_ranges


def input_output_match(input_rows, input_cols, output_rows, output_cols):
    if len(input_rows) != len(output_rows) or len(input_cols) != len(output_cols):
        return False
    for i, row in enumerate(input_rows):
        if len(output_rows[i]) != len(row):
            return False
        elif len(row) == 2:
            if (row[1] - row[0]) != (output_rows[i][1] - output_rows[i][0]):
                return False
    for i, col in enumerate(input_cols):
        if len(output_cols[i]) != len(col):
            return False
        elif len(col) == 2:
            if (col[1] - col[0]) != (output_cols[i][1] - output_cols[i][0]):
                return False
    return True


def cpy_excel_main(input_config, output_config):
    workbook = open_excel(input_config[0][1])
    sheet = open_sheet(workbook)
    # e.read_excel(filename)
    rows = get_row_name(input_config[2][1])
    cols = get_col_name(input_config[3][1])
    rows1 = get_row_name(output_config[2][1])
    cols1 = get_col_name(output_config[3][1])
    print(rows, cols, rows1, cols1)
    if input_output_match(rows, cols, rows1, cols1):
        print("config is correct")
        data = read_excel(sheet, rows, cols)
        write_excel_xlsx(output_config[0][1], data, rows, cols, output_config[1][1])
    else:
        print("input output config not match")

